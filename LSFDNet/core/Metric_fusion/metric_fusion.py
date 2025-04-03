import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import statistics
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
    # 文件不存在，创建新的 Workbook
        workbook = Workbook()

    # 获取或创建一个工作表
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # 在指定列中插入数据
    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i+1)]
        cell.value = value

    # 保存文件
    workbook.save(excel_name)

def write_excel_any(excel_name='metric.xlsx', worksheet_name='VIF', row_index=0, column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        # 文件不存在，创建新的 Workbook
        workbook = Workbook()

    # 获取或创建一个工作表
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # 在指定的单元格中插入数据
    cell = worksheet.cell(row=row_index + 1, column=column_index + 1)
    cell.value = data

    # 保存文件
    workbook.save(excel_name)

def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)

    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM

if __name__ == '__main__':
    with_mean = True
    dataroot = r'/home/cc/LY/UMF-CMGR-main/Evaluation/Metric_fusion/datasets'
    results_root = '/home/cc/LY/UMF-CMGR-main/Evaluation/Metric_fusion/Results'
    dataset = 'TNO'
    ir_dir = os.path.join(dataroot, dataset, 'ir')
    vi_dir = os.path.join(dataroot, dataset, 'vi')
    f_dir = os.path.join(results_root, dataset)
    save_dir = '/home/cc/LY/UMF-CMGR-main/Evaluation/Metric_fusion/metric'
    os.makedirs(save_dir, exist_ok=True)

    metric_save_name = os.path.join(save_dir, 'metric_result.xlsx')
    filelist = natsorted(os.listdir(ir_dir))

    Method_list = ['GTF', 'DIDFuse', 'RFN-Nest', 'FusionGAN', 'TarDAL', 'SeAFusion', 'SwinFusion', 'U2Fusion', 'PSF', 'UMF-CMGR_RGB','UMF-CMGR-semantic_1000']
    for i, Method in enumerate(Method_list):
        EN_list = []
        MI_list = []
        SF_list = []
        AG_list = []
        SD_list = []
        CC_list = []
        SCD_list = []
        VIF_list = []
        MSE_list = []
        PSNR_list = []
        Qabf_list = []
        Nabf_list = []
        SSIM_list = []
        MS_SSIM_list = []
        filename_list = ['','EN', 'MI', 'SF', 'AG', 'SD', 'CC', 'SCD', 'VIF', 'MSE', 'PSNR', 'Qabf', 'Nabf', 'SSIM', 'MS_SSIM']
        sub_f_dir = os.path.join(f_dir, Method)
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(sub_f_dir, item)
            print(ir_name, vi_name, f_name)
            EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name,
                                                                                                    f_name)
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            AG_list.append(AG)
            SD_list.append(SD)
            CC_list.append(CC)
            SCD_list.append(SCD)
            VIF_list.append(VIF)
            MSE_list.append(MSE)
            PSNR_list.append(PSNR)
            Qabf_list.append(Qabf)
            Nabf_list.append(Nabf)
            SSIM_list.append(SSIM)
            MS_SSIM_list.append(MS_SSIM)
            #filename_list.append(item)
            eval_bar.set_description("{} | {}".format(Method, item))
        if with_mean:
            # 添加均值
            EN_list.insert(0, np.mean(EN_list))
            MI_list.insert(0, np.mean(MI_list))
            SF_list.insert(0, np.mean(SF_list))
            AG_list.insert(0, np.mean(AG_list))
            SD_list.insert(0, np.mean(SD_list))
            CC_list.insert(0, np.mean(CC_list))
            SCD_list.insert(0, np.mean(SCD_list))
            VIF_list.insert(0, np.mean(VIF_list))
            MSE_list.insert(0, np.mean(MSE_list))
            PSNR_list.insert(0, np.mean(PSNR_list))
            Qabf_list.insert(0, np.mean(Qabf_list))
            Nabf_list.insert(0, np.mean(Nabf_list))
            SSIM_list.insert(0, np.mean(SSIM_list))
            MS_SSIM_list.insert(0, np.mean(MS_SSIM_list))


            # EN_list.append(np.mean(EN_list))
            # MI_list.append(np.mean(MI_list))
            # SF_list.append(np.mean(SF_list))
            # AG_list.append(np.mean(AG_list))
            # SD_list.append(np.mean(SD_list))
            # CC_list.append(np.mean(CC_list))
            # SCD_list.append(np.mean(SCD_list))
            # VIF_list.append(np.mean(VIF_list))
            # MSE_list.append(np.mean(MSE_list))
            # PSNR_list.append(np.mean(PSNR_list))
            # Qabf_list.append(np.mean(Qabf_list))
            # Nabf_list.append(np.mean(Nabf_list))
            # SSIM_list.append(np.mean(SSIM_list))
            # MS_SSIM_list.append(np.mean(MS_SSIM_list))

            #filename_list.append('mean')

        ## 保留三位小数
        EN_list = [round(x, 3) for x in EN_list]
        MI_list = [round(x, 3) for x in MI_list]
        SF_list = [round(x, 3) for x in SF_list]
        AG_list = [round(x, 3) for x in AG_list]
        SD_list = [round(x, 3) for x in SD_list]
        CC_list = [round(x, 3) for x in CC_list]
        SCD_list = [round(x, 3) for x in SCD_list]
        VIF_list = [round(x, 3) for x in VIF_list]
        MSE_list = [round(x, 3) for x in MSE_list]
        PSNR_list = [round(x, 3) for x in PSNR_list]
        Qabf_list = [round(x, 3) for x in Qabf_list]
        Nabf_list = [round(x, 3) for x in Nabf_list]
        SSIM_list = [round(x, 3) for x in SSIM_list]
        MS_SSIM_list = [round(x, 3) for x in MS_SSIM_list]

        EN_list.insert(0, '{}'.format(Method))
        MI_list.insert(0, '{}'.format(Method))
        SF_list.insert(0, '{}'.format(Method))
        AG_list.insert(0, '{}'.format(Method))
        SD_list.insert(0, '{}'.format(Method))
        CC_list.insert(0, '{}'.format(Method))
        SCD_list.insert(0, '{}'.format(Method))
        VIF_list.insert(0, '{}'.format(Method))
        MSE_list.insert(0, '{}'.format(Method))
        PSNR_list.insert(0, '{}'.format(Method))
        Qabf_list.insert(0, '{}'.format(Method))
        Nabf_list.insert(0, '{}'.format(Method))
        SSIM_list.insert(0, '{}'.format(Method))
        MS_SSIM_list.insert(0, '{}'.format(Method))
        if i == 0:
            write_excel(metric_save_name, "VIF", 0, filename_list)
        write_excel_any(metric_save_name, "VIF", 0, i+1, EN_list[0])
        write_excel_any(metric_save_name, "VIF", 1, i+1, EN_list[1])
        write_excel_any(metric_save_name, "VIF", 2, i+1, MI_list[1])
        write_excel_any(metric_save_name, "VIF", 3, i+1, SF_list[1])
        write_excel_any(metric_save_name, "VIF", 4, i+1, AG_list[1])
        write_excel_any(metric_save_name, "VIF", 5, i+1, SD_list[1])
        write_excel_any(metric_save_name, "VIF", 6, i+1, CC_list[1])
        write_excel_any(metric_save_name, "VIF", 7, i+1, SCD_list[1])
        write_excel_any(metric_save_name, "VIF", 8, i+1, VIF_list[1])
        write_excel_any(metric_save_name, "VIF", 9, i+1, MSE_list[1])
        write_excel_any(metric_save_name, "VIF", 10, i+1, PSNR_list[1])
        write_excel_any(metric_save_name, "VIF", 11, i+1, Qabf_list[1])
        write_excel_any(metric_save_name, "VIF", 12, i+1, Nabf_list[1])
        write_excel_any(metric_save_name, "VIF", 13, i+1, SSIM_list[1])
        write_excel_any(metric_save_name, "VIF", 14, i+1, MS_SSIM_list[1])




